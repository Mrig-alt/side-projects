from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "Daily Log"

# ── Colours ──────────────────────────────────────────────────────────────────
BLACK  = "FF000000"
ORANGE = "FFFFA500"
WHITE  = "FFFFFFFF"
DGREY  = "FF1A1A1A"   # slightly lighter black for data rows

black_fill  = PatternFill("solid", fgColor=BLACK)
dgrey_fill  = PatternFill("solid", fgColor=DGREY)
orange_font = Font(name="Calibri", bold=True, color=ORANGE, size=11)
white_font  = Font(name="Calibri", italic=True, color=WHITE, size=11)
note_font   = Font(name="Calibri", italic=True, color="FF888888", size=10)

center = Alignment(horizontal="center", vertical="center")
left   = Alignment(horizontal="left",   vertical="center")

# ── Column widths ─────────────────────────────────────────────────────────────
col_widths = {"A": 16, "B": 28, "C": 16, "D": 18, "E": 10}
for col, w in col_widths.items():
    ws.column_dimensions[col].width = w

ws.row_dimensions[1].height = 24

# ── Header row (row 1) ────────────────────────────────────────────────────────
headers = ["Date", "Task", "Time Selected", "Status", "Week No"]
for i, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=i, value=h)
    cell.fill = black_fill
    cell.font = orange_font
    cell.alignment = center

# ── Sample data rows ──────────────────────────────────────────────────────────
rows = [
    ("19 Mar 2026", "Meryl meeting",    "08:14", "pending",       12),
    ("19 Mar 2026", "Odd jobs apply",   "08:14", "pending",       12),
    ("19 Mar 2026", "LinkedIn tracker", "08:15", "✅ Done",       12),
    ("20 Mar 2026", "Python with Jose", "08:09", "❌ Not done",   12),
]

for r, (date, task, time_, status, week) in enumerate(rows, 2):
    ws.row_dimensions[r].height = 20
    for c, val in enumerate([date, task, time_, status, week], 1):
        cell = ws.cell(row=r, column=c, value=val)
        cell.fill = black_fill
        cell.font = white_font
        cell.alignment = center if c != 2 else left

# ── Blank spacer row 6 ────────────────────────────────────────────────────────
for c in range(1, 6):
    ws.cell(row=6, column=c).fill = black_fill

# ── Note row 7 ────────────────────────────────────────────────────────────────
ws.merge_cells("A7:E7")
note = ws.cell(row=7, column=1,
               value="↑ Sample rows — Make will overwrite these automatically")
note.font = note_font
note.alignment = center
note.fill = black_fill
ws.row_dimensions[7].height = 18

# ── Fill remaining rows black (8-50) ─────────────────────────────────────────
for r in range(8, 51):
    ws.row_dimensions[r].height = 18
    for c in range(1, 6):
        ws.cell(row=r, column=c).fill = black_fill

# ── Freeze header row ─────────────────────────────────────────────────────────
ws.freeze_panes = "A2"

out = "/home/user/side-projects/whatsapp_task_tracker.xlsx"
wb.save(out)
print(f"Saved → {out}")
