"""
excel_store.py — Full ADHD progress log written to tracker.xlsx.

Two sheets:

  "Daily Log"   — append-only, one row per (date × task).
                  Captures every step: shown → planned → completed.
    Columns: Date | Task Name | Type | Shown | Planned | Completed | Skipped | Day Score (%) | Streak

  "Task Stats"  — one row per task, overwritten after each evening poll.
                  Shows development over time.
    Columns: Task Name | Type | Days Shown | Days Planned | Days Completed |
             Days Missed | Completion Rate (%) | Current Streak | Best Streak | Last Active

Usage flow:
  1. Morning poll sent      → log_morning_poll_sent()    writes Shown=True rows
  2. Morning vote received  → write_morning_selections() updates Planned / Skipped
  3. Evening vote received  → write_evening_completions() updates Completed / Skipped / Score / Streak
                           → refresh_task_stats()        rewrites Task Stats sheet
"""

import threading
from datetime import date as date_type
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from config import EXCEL_FILE

# ── Sheet names & headers ─────────────────────────────────────────────────────

SHEET_LOG   = "Daily Log"
SHEET_STATS = "Task Stats"

LOG_HEADERS = [
    "Date", "Task Name", "Type",
    "Shown", "Planned", "Completed", "Skipped",
    "Day Score (%)", "Streak",
]

STATS_HEADERS = [
    "Task Name", "Type",
    "Days Shown", "Days Planned", "Days Completed", "Days Missed",
    "Completion Rate (%)", "Current Streak", "Best Streak", "Last Active",
]

# Column indices for Daily Log (1-based)
L_DATE      = 1
L_NAME      = 2
L_TYPE      = 3
L_SHOWN     = 4
L_PLANNED   = 5
L_COMPLETED = 6
L_SKIPPED   = 7
L_SCORE     = 8
L_STREAK    = 9

_lock = threading.Lock()

# ── Colours ───────────────────────────────────────────────────────────────────

_BLUE  = "4472C4"
_GREEN = "70AD47"

# ── Workbook helpers ──────────────────────────────────────────────────────────

def _open_or_create() -> openpyxl.Workbook:
    if EXCEL_FILE.exists():
        return openpyxl.load_workbook(EXCEL_FILE)
    wb = openpyxl.Workbook()
    # rename default sheet to Daily Log
    ws = wb.active
    ws.title = SHEET_LOG
    _write_headers(ws, LOG_HEADERS, _BLUE)
    _set_widths(ws, [12, 40, 12, 8, 10, 12, 10, 14, 8])
    # create Task Stats sheet
    ws2 = wb.create_sheet(SHEET_STATS)
    _write_headers(ws2, STATS_HEADERS, _GREEN)
    _set_widths(ws2, [40, 12, 12, 14, 16, 13, 18, 15, 12, 14])
    wb.save(EXCEL_FILE)
    return wb


def _write_headers(ws, headers: list[str], colour: str) -> None:
    fill = PatternFill("solid", fgColor=colour)
    font = Font(bold=True, color="FFFFFF")
    for col, title in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")


def _set_widths(ws, widths: list[int]) -> None:
    for col, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = w


def _ensure_sheets(wb: openpyxl.Workbook) -> tuple:
    """Return (log_sheet, stats_sheet), creating them if missing."""
    if SHEET_LOG not in wb.sheetnames:
        ws = wb.create_sheet(SHEET_LOG, 0)
        _write_headers(ws, LOG_HEADERS, _BLUE)
        _set_widths(ws, [12, 40, 12, 8, 10, 12, 10, 14, 8])
    if SHEET_STATS not in wb.sheetnames:
        ws2 = wb.create_sheet(SHEET_STATS)
        _write_headers(ws2, STATS_HEADERS, _GREEN)
        _set_widths(ws2, [40, 12, 12, 14, 16, 13, 18, 15, 12, 14])
    return wb[SHEET_LOG], wb[SHEET_STATS]


# ── Row lookup in Daily Log ───────────────────────────────────────────────────

def _find_log_rows(ws, day: str) -> dict[str, int]:
    """Return {task_name: row_number} for all rows matching the given date."""
    result = {}
    for row in ws.iter_rows(min_row=2):
        if str(row[L_DATE - 1].value) == day:
            name = row[L_NAME - 1].value
            if name:
                result[str(name)] = row[L_DATE - 1].row
    return result


# ── Step 1: Morning poll sent ─────────────────────────────────────────────────

def log_morning_poll_sent(day: str, task_names: list[str], task_types: list[str]) -> None:
    """
    Called the moment the morning poll is dispatched.
    Creates one row per task with Shown=True. All other columns left blank
    until votes come in — this way every sent poll is on record even if
    the user never votes.
    """
    with _lock:
        wb = _open_or_create()
        ws, _ = _ensure_sheets(wb)
        existing = _find_log_rows(ws, day)

        for name, ttype in zip(task_names, task_types):
            if name not in existing:
                ws.append([day, name, ttype, True, None, None, None, None, None])

        wb.save(EXCEL_FILE)


# ── Step 2: Morning vote recorded ─────────────────────────────────────────────

def write_morning_selections(
    day: str,
    task_names: list[str],
    task_types: list[str],
    planned_indices: list[int],
) -> None:
    """
    Called when the user submits their morning plan.
    Updates Planned and Skipped. Creates rows for any tasks not yet logged
    (handles case where log_morning_poll_sent was not called, e.g. first run).
    """
    planned_set = set(planned_indices)

    with _lock:
        wb = _open_or_create()
        ws, _ = _ensure_sheets(wb)
        existing = _find_log_rows(ws, day)

        for i, (name, ttype) in enumerate(zip(task_names, task_types)):
            planned = i in planned_set
            skipped = not planned

            if name in existing:
                row_num = existing[name]
                ws.cell(row=row_num, column=L_PLANNED).value = planned
                ws.cell(row=row_num, column=L_SKIPPED).value = skipped
            else:
                ws.append([day, name, ttype, True, planned, None, skipped, None, None])

        wb.save(EXCEL_FILE)


# ── Step 3: Evening vote recorded ─────────────────────────────────────────────

def write_evening_completions(
    day: str,
    task_names: list[str],
    task_types: list[str],
    completed_indices: list[int],
    score: int,
    streaks: dict[str, int],
) -> None:
    """
    Called when the user submits their evening completion check.
    Updates Completed, re-evaluates Skipped, fills Day Score and Streak.
    streaks is a dict of {task_name: current_streak_value}.
    """
    completed_set = set(completed_indices)

    with _lock:
        wb = _open_or_create()
        ws, _ = _ensure_sheets(wb)
        existing = _find_log_rows(ws, day)

        for i, (name, ttype) in enumerate(zip(task_names, task_types)):
            completed = i in completed_set
            streak = streaks.get(name)

            if name in existing:
                row_num = existing[name]
                planned = ws.cell(row=row_num, column=L_PLANNED).value
                skipped = not planned and not completed
                ws.cell(row=row_num, column=L_COMPLETED).value = completed
                ws.cell(row=row_num, column=L_SKIPPED).value   = skipped
                ws.cell(row=row_num, column=L_SCORE).value     = score
                ws.cell(row=row_num, column=L_STREAK).value    = streak
            else:
                # Morning poll was skipped entirely — create row now
                skipped = not completed
                ws.append([day, name, ttype, None, None, completed, skipped, score, streak])

        # Back-fill score on any rows that already existed for this day
        for row_num in existing.values():
            ws.cell(row=row_num, column=L_SCORE).value = score

        wb.save(EXCEL_FILE)


# ── Task Stats sheet ──────────────────────────────────────────────────────────

def refresh_task_stats(task_objects: list[dict]) -> None:
    """
    Rewrite the Task Stats sheet from the current task objects.
    Called after every evening vote so the sheet always reflects
    the latest streaks and completion counts.
    """
    with _lock:
        wb = _open_or_create()
        _, ws = _ensure_sheets(wb)

        # Clear existing data rows (keep header)
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.value = None

        # Find last non-empty row after clearing, then trim
        row_num = 2
        for task in task_objects:
            name        = task.get("name", "")
            ttype       = task.get("type", "")
            completed   = task.get("completed", 0)
            missed      = task.get("missed", 0)
            shown       = completed + missed
            planned     = shown  # every shown task was planned by definition in tasks.json mode
            rate        = round(completed / shown * 100) if shown else 0
            cur_streak  = task.get("current_streak", 0)
            best_streak = task.get("best_streak", 0)
            last_active = task.get("last_active", "")

            ws.cell(row=row_num, column=1,  value=name)
            ws.cell(row=row_num, column=2,  value=ttype)
            ws.cell(row=row_num, column=3,  value=shown)
            ws.cell(row=row_num, column=4,  value=planned)
            ws.cell(row=row_num, column=5,  value=completed)
            ws.cell(row=row_num, column=6,  value=missed)
            ws.cell(row=row_num, column=7,  value=rate)
            ws.cell(row=row_num, column=8,  value=cur_streak)
            ws.cell(row=row_num, column=9,  value=best_streak)
            ws.cell(row=row_num, column=10, value=last_active)
            row_num += 1

        wb.save(EXCEL_FILE)
