"""
excel_store.py — Persistent Excel log for daily task selections and completions.

File: tracker.xlsx  (sync this file to OneDrive so Make.com can read it)

Sheet "Daily Log" — one row per (date × task):
  Date | Task Index | Task Name | Planned | Completed | Skipped | Day Score (%)

Planned   = True  → selected in morning poll
Completed = True  → marked done in evening poll
Skipped   = True  → task existed but was not planned AND not completed
Day Score = overall % for the day (same value on every row for that day)
"""

import threading
from datetime import date as date_type
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

EXCEL_FILE = Path(__file__).parent / "tracker.xlsx"

SHEET_LOG = "Daily Log"
HEADERS = ["Date", "Task Index", "Task Name", "Planned", "Completed", "Skipped", "Day Score (%)"]

# Column indices (1-based)
COL_DATE       = 1
COL_TASK_IDX   = 2
COL_TASK_NAME  = 3
COL_PLANNED    = 4
COL_COMPLETED  = 5
COL_SKIPPED    = 6
COL_SCORE      = 7

_lock = threading.Lock()


# ── Workbook helpers ──────────────────────────────────────────────────────────

def _open_or_create() -> openpyxl.Workbook:
    if EXCEL_FILE.exists():
        return openpyxl.load_workbook(EXCEL_FILE)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SHEET_LOG
    _write_header_row(ws)
    _set_column_widths(ws)
    wb.save(EXCEL_FILE)
    return wb


def _write_header_row(ws) -> None:
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF")
    for col, title in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")


def _set_column_widths(ws) -> None:
    widths = [12, 12, 45, 10, 12, 10, 14]
    for col, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = w


def _get_sheet(wb: openpyxl.Workbook):
    if SHEET_LOG in wb.sheetnames:
        return wb[SHEET_LOG]
    ws = wb.active
    ws.title = SHEET_LOG
    _write_header_row(ws)
    _set_column_widths(ws)
    return ws


# ── Row lookup ────────────────────────────────────────────────────────────────

def _find_rows_for_day(ws, day: str) -> dict[int, int]:
    """Return {task_index: row_number} for all rows matching the given date."""
    result = {}
    for row in ws.iter_rows(min_row=2):
        if str(row[COL_DATE - 1].value) == day:
            task_idx = row[COL_TASK_IDX - 1].value
            if task_idx is not None:
                result[int(task_idx)] = row[COL_DATE - 1].row
    return result


# ── Public API ────────────────────────────────────────────────────────────────

def write_morning_selections(day: str, tasks: list[str], planned_indices: list[int]) -> None:
    """
    Called when user answers the morning poll.
    Writes one row per task: Planned=True if selected, Skipped=True if not.
    Completed is left blank until the evening poll.
    """
    planned_set = set(planned_indices)

    with _lock:
        wb = _open_or_create()
        ws = _get_sheet(wb)
        existing = _find_rows_for_day(ws, day)

        for i, task_name in enumerate(tasks):
            planned  = i in planned_set
            skipped  = not planned  # will be re-evaluated after evening if needed

            if i in existing:
                row_num = existing[i]
                ws.cell(row=row_num, column=COL_PLANNED).value  = planned
                ws.cell(row=row_num, column=COL_SKIPPED).value  = skipped
            else:
                ws.append([day, i, task_name, planned, None, skipped, None])

        wb.save(EXCEL_FILE)


def write_evening_completions(
    day: str,
    tasks: list[str],
    completed_indices: list[int],
    score: int,
) -> None:
    """
    Called when user answers the evening poll.
    Updates Completed, re-evaluates Skipped, and fills Day Score for all rows.
    Creates rows for any tasks that don't have a morning entry (morning skipped).
    """
    completed_set = set(completed_indices)

    with _lock:
        wb = _open_or_create()
        ws = _get_sheet(wb)
        existing = _find_rows_for_day(ws, day)

        for i, task_name in enumerate(tasks):
            completed = i in completed_set

            if i in existing:
                row_num = existing[i]
                planned = ws.cell(row=row_num, column=COL_PLANNED).value
                skipped = not planned and not completed
                ws.cell(row=row_num, column=COL_COMPLETED).value = completed
                ws.cell(row=row_num, column=COL_SKIPPED).value   = skipped
                ws.cell(row=row_num, column=COL_SCORE).value     = score
            else:
                # Morning poll was skipped — create row now
                skipped = not completed
                ws.append([day, i, task_name, None, completed, skipped, score])

        # Back-fill score on any rows already written for this day
        for row_num in existing.values():
            ws.cell(row=row_num, column=COL_SCORE).value = score

        wb.save(EXCEL_FILE)


def get_planned_indices_for_date(day: str) -> list[int] | None:
    """
    Read back planned task indices from Excel for a given day.
    Returns None if no morning data exists for that day.
    """
    if not EXCEL_FILE.exists():
        return None

    with _lock:
        wb = openpyxl.load_workbook(EXCEL_FILE, read_only=True)
        ws = wb[SHEET_LOG] if SHEET_LOG in wb.sheetnames else wb.active

        found_day = False
        planned = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if str(row[COL_DATE - 1]) == day:
                found_day = True
                if row[COL_PLANNED - 1] is True:
                    planned.append(int(row[COL_TASK_IDX - 1]))

        wb.close()
        return planned if found_day else None
